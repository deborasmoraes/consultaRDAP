import openpyxl
import requests
from tqdm import tqdm
import time

# Carregar a planilha de domínios
dominios_workbook = openpyxl.load_workbook('dominios.xlsx')
dominios_sheet = dominios_workbook.active
dominios_coluna = dominios_sheet['A']

# Carregar a planilha de CPF/CNPJ
dados_workbook = openpyxl.load_workbook('dadosCPF.xlsx')
dados_sheet = dados_workbook['planosp']
cpf_cnpj_coluna = [cell.value for cell in dados_sheet['C'][1:]]

# Criar a planilha de resultados
resultados_workbook = openpyxl.Workbook()
resultados_sheet = resultados_workbook.active
resultados_sheet['A1'] = 'Domínio'
resultados_sheet['B1'] = 'CPF/CNPJ'
resultados_sheet['C1'] = 'Localizado'

# URL base para consulta RDAP
base_url = 'https://rdap.registro.br/domain/'

s = requests.Session()

for i, dominio_celula in tqdm(enumerate(dominios_coluna, start=1), desc='Consultando Domínios', total=len(dominios_coluna)):
    dominio = dominio_celula.value
    print(f"Consultando domínio: {dominio}")

    url = base_url + dominio
    print(f"URL da consulta: {url}")

    response = s.get(url, verify=False)
    print(f"Resposta da consulta: {response.text}")
    if response.status_code == 200:
        data = response.json()
        entities = data.get('entities', [])
        cpf_cnpj = None
        localizado = ''

        for entity in entities:
            identifiers = entity.get('publicIds', [])
            for identifier in identifiers:
                if identifier['type'] == 'cnpj' or identifier['type'] == 'cpf':
                    cpf_cnpj = identifier['identifier']
                    if cpf_cnpj in cpf_cnpj_coluna:
                        localizado = 'Sim'
                        break

        resultados_sheet.cell(row=i+1, column=1, value=dominio)
        resultados_sheet.cell(row=i+1, column=2, value=cpf_cnpj)
        resultados_sheet.cell(row=i+1, column=3, value=localizado)

    # Aguarda 7 minutos a cada 29 solicitações
    if i % 29 == 0:
        print("Aguardando 7 minutos...")
        time.sleep(420)  # 7 minutos = 420 segundos

# Salvar a planilha de resultados
resultados_workbook.save('resultados.xlsx')
